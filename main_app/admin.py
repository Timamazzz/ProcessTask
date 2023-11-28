import openpyxl
from django.contrib import admin
from django.http import HttpResponse

from .enums import ServiceType, ProcessStatus, LifeSituationName
from .models import CustomUser, LifeSituation, Process, Service, Organization, ProcessGroup
from django.contrib.auth.admin import UserAdmin


class OrganizationAdmin(admin.ModelAdmin):
    list_display = ('name', 'code')

    actions = ['export_to_excel']

    def export_to_excel(self, request, queryset):
        # Create a new Excel workbook
        wb = openpyxl.Workbook()

        # Populate each worksheet with data from the selected Organization
        for organization in queryset:
            ws = wb.create_sheet(title=organization.code)  # Use organization code as sheet name

            # Write headers and merge cells
            headers = [
                ('Жизненная ситуация', 'A', 'C'),
                ('Услуга', 'D', 'H'),
                ('Процесс', 'I', 'X'),
            ]

            for header, start_column, end_column in headers:
                ws[f'{start_column}1'] = header
                ws.merge_cells(f'{start_column}1:{end_column}1')

            # Write the common headers in the second row
            common_headers = [
                'Идентификатор', 'Название', 'Пользователь',
                'Идентификатор', 'Название', 'Тип услуги', 'Регулирующий акт', 'Пользователь',
                'Идентификатор', 'Название', 'Статус', 'Внутренний клиент', 'Внешний клиент',
                'Ответственный орган', 'Отдел', 'Цифровой формат', 'Не цифровой формат',
                'Ссылка на цифровой формат', 'Ценность для клиента', 'Данные на входе', 'Данные на выходе',
                'Связанные процессы', 'Группа', 'Пользователь',
            ]

            for col_num, header in enumerate(common_headers, start=1):
                ws.cell(row=2, column=col_num, value=header)

                # Populate the worksheet with data for the current organization
                row_num = 3  # Start from the third row

                life_situations = LifeSituation.objects.filter(user__organization=organization)

                for life_situation in life_situations:
                    ws.cell(row=row_num, column=1, value=life_situation.identifier)
                    ws.cell(row=row_num, column=2, value=LifeSituationName[life_situation.name].value)
                    ws.cell(row=row_num, column=3, value=life_situation.user.email)

                    start_life_situation_row = row_num

                    # Fetch related services for the current life situation
                    services = Service.objects.filter(lifesituation=life_situation)

                    for service_index, service in enumerate(services):
                        ws.cell(row=row_num, column=4, value=service.identifier)
                        ws.cell(row=row_num, column=5, value=service.name)
                        ws.cell(row=row_num, column=6, value=ServiceType[service.service_type].value)
                        ws.cell(row=row_num, column=7, value=service.regulating_act)
                        ws.cell(row=row_num, column=8, value=service.user.email)

                        start_service_row = row_num

                        # Fetch related processes for the current service
                        processes = Process.objects.filter(service=service)

                        for process_index, process in enumerate(processes):
                            ws.cell(row=row_num, column=9, value=process.identifier)
                            ws.cell(row=row_num, column=10, value=process.name)
                            ws.cell(row=row_num, column=11, value=ProcessStatus[process.status].value)
                            ws.cell(row=row_num, column=12, value='Да' if process.is_internal_client else 'Нет')
                            ws.cell(row=row_num, column=13, value='Да' if process.is_external_client else 'Нет')
                            ws.cell(row=row_num, column=14, value=process.responsible_authority)
                            ws.cell(row=row_num, column=15, value=process.department)
                            ws.cell(row=row_num, column=16, value='Да' if process.is_digital_format else 'Нет')
                            ws.cell(row=row_num, column=17, value='Да' if process.is_non_digital_format else 'Нет')
                            ws.cell(row=row_num, column=18, value=process.digital_format_link)
                            ws.cell(row=row_num, column=19, value=process.client_value)
                            ws.cell(row=row_num, column=20, value=process.input_data)
                            ws.cell(row=row_num, column=21, value=process.output_data)
                            ws.cell(row=row_num, column=22, value=process.related_processes)
                            try:
                                ws.cell(row=row_num, column=23, value=process.group.name)
                            except KeyError:
                                ws.cell(row=row_num, column=23, value='')
                            ws.cell(row=row_num, column=24, value=process.user.email)

                            if process_index < len(processes) - 1:
                                row_num += 1

                        # Merge cells for the service columns
                        if len(processes) > 0:
                            for col_num in range(4, 9):
                                ws.merge_cells(start_row=start_service_row, start_column=col_num,
                                               end_row=row_num, end_column=col_num)
                        if service_index < len(services) - 1:
                            row_num += 1

                    # Merge cells for the life situation columns
                    if len(services) > 0:
                        for col_num in range(1, 4):
                            ws.merge_cells(start_row=start_life_situation_row, start_column=col_num,
                                           end_row=row_num, end_column=col_num)
                    row_num += 1

        # Remove the default sheet created and save the response
        wb.remove(wb.active)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'
        wb.save(response)
        return response

    export_to_excel.short_description = "Экспорт данных выбранных организаций в Excel"


class CustomUserAdmin(UserAdmin):
    list_display = ('email', 'first_name', 'last_name', 'patronymic', 'organization')
    search_fields = ('email', 'first_name', 'last_name', 'patronymic', 'organization__name')
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'groups', 'date_joined', 'organization')
    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        ('Personal Info', {'fields': ('first_name', 'last_name', 'patronymic', 'organization')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Important dates', {'fields': ('last_login', 'date_joined')}),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'password1', 'password2')}
         ),
    )
    ordering = ('email',)


class LifeSituationAdmin(admin.ModelAdmin):
    list_display = ('name', 'user')
    search_fields = ('name', 'user__email')
    list_filter = ('user', 'name')


class ServiceAdmin(admin.ModelAdmin):
    list_display = ('name', 'service_type', 'regulating_act', 'lifesituation', 'user')
    search_fields = ('name', 'service_type', 'regulating_act', 'user__email')
    list_filter = ('service_type', 'lifesituation', 'user')


class ProcessGroupAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)


class ProcessAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'service', 'status', 'is_internal_client', 'is_external_client', 'responsible_authority', 'department',
        'is_digital_format', 'is_non_digital_format', 'group')
    search_fields = ('name', 'service__name', 'responsible_authority', 'department')
    list_filter = ('status', 'is_internal_client', 'is_external_client', 'is_digital_format', 'is_non_digital_format',
                   'group', 'service', 'user__organization')

    fieldsets = (
        (None, {'fields': ('name', 'service', 'status', 'identifier')}),
        ('Responsibility', {'fields': ('responsible_authority', 'department')}),
        ('Digital Format', {'fields': ('is_digital_format', 'is_non_digital_format', 'digital_format_link')}),
        ('Data', {'fields': ('client_value', 'input_data', 'output_data', 'related_processes', 'group')}),
    )


admin.site.register(Organization, OrganizationAdmin)
admin.site.register(CustomUser, CustomUserAdmin)
admin.site.register(LifeSituation, LifeSituationAdmin)
admin.site.register(Service, ServiceAdmin)
admin.site.register(ProcessGroup, ProcessGroupAdmin)
admin.site.register(Process, ProcessAdmin)
